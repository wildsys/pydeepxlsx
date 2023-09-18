#!/usr/bin/env python
# coding: utf-8

# Imports
from copy import copy
from slugify import slugify
from openpyxl.utils.cell import get_column_letter
import openpyxl
import os
import re


class PyDeepXLSX:
    wbtpl = None                        #: Template file as openpyxl workbook object
    wb = None                           #: Result file as openpyxl workbook object
    copy_sheet_attr_to_preserve = []    #: List of sheet attributes to preserve in copy
    copy_column_attr_to_preserve = []   #: List of column attributes to preserve in copy

    def __init__(self, templated_file):
        """
        `templated_file` is the path to the XLSX to use as template
        """

        assert os.path.isfile(templated_file), f'{templated_file} does not exists or is not a file'
        self.wbtpl = openpyxl.load_workbook(filename=templated_file)
        self.wb = openpyxl.Workbook()
        self.wb.remove_sheet(self.wb.active)
        self.copy_sheet_attr_to_preserve = [
            'sheet_format',
            'sheet_properties',
            'merged_cells',
            'page_margins',
            'freeze_panes'
        ]
        self.copy_column_attr_to_preserve = [
            'min',
            'max',
            'width',
            'hidden'
        ]

    def copy_as_it(self, sheetname, rowstop=None):
        """
        Copy a sheet as it (static contents or first `rowstop` rows)
        """

        assert rowstop is None or isinstance(rowstop, (int, )), '{rowstop} can only be an integer'

        src = self.wbtpl.get_sheet_by_name(sheetname)
        dst = self.wb.create_sheet(title=src.title)

        # Copy all sheet attributes
        for attname in self.copy_sheet_attr_to_preserve:
            setattr(dst, attname, copy(getattr(src, attname)))

        # Copy row styles
        dst.row_dimensions = copy(src.row_dimensions)

        # Copy column styles
        if src.sheet_format.defaultColWidth is not None:
            dst.sheet_format.defaultColWidth = copy(src.sheet_format.defaultColWidth)
        for cdk in src.column_dimensions.keys():
            for attname in self.copy_column_attr_to_preserve:
                copy_data = copy(getattr(src.column_dimensions[cdk], attname))
                setattr(dst.column_dimensions[cdk], attname, copy_data)

        # Copy rows content
        count_rows = min(rowstop, src.max_row) if rowstop is not None else src.max_row
        for row in range(count_rows + 1):
            for cell in range(src.max_column + 1):
                sc = src.cell(row + 1, cell + 1)
                c = dst.cell(row + 1, cell + 1, value=sc.value)
                c.font = copy(sc.font)
                c.border = copy(sc.border)
                c.fill = copy(sc.fill)
                c.number_format = copy(sc.number_format)
                c.protection = copy(sc.protection)
                c.alignment = copy(sc.alignment)

    def get_column_names(self, sheetname, headers=0):
        """
        Returns the list of column names slugified in the `sheetname`

        Column name is given by the `headers` row (0-based count)
        """
        output = []
        ws = self.wbtpl.get_sheet_by_name(sheetname)
        for idx in range(ws.max_column):
            cell = ws.cell((headers + 1), (idx + 1))
            value = cell.value
            if value:
                value = slugify(value, separator='_')
            output.append(value)
        return output

    def get_template_line(self, sheetname, headers=0):
        """
        Returns the template of one line

        Line taken is the first aftet the `headers` row (0-based count)
        """
        output = []
        ws = self.wbtpl.get_sheet_by_name(sheetname)
        for idx in range(ws.max_column):
            cell = ws.cell((headers + 1), (idx + 1))
            output.append(cell)
        return output

    def cell_copy(self, sheetname, row, cell, headers, value):
        """
        Copy a cell from the `sheetname` source to destionation sheet.

        `value` applies at position `row` x `cell` if not a formula.

        `headers` is used to compute the first data line (= `headers` + 1) to use as source for all rows (0-based).
        """
        src = self.wbtpl.get_sheet_by_name(sheetname)
        dst = self.wb.get_sheet_by_name(sheetname)
        sc = src.cell((headers + 2), cell)
        if isinstance(sc.value, (str, )) and sc.value.startswith('='):
            value = sc.value
            matches = list(set(re.findall(r'([A-Z]+\d+)', value)))
            for m in matches:
                explore = re.search(r'^([A-Z]+)(\d+)$', m)
                updated_row = int(explore[2]) - (headers + 2) + row
                updated_cell = f'{explore[1]}{updated_row}'
                value = value.replace(m, updated_cell)
        c = dst.cell(row, cell, value=value)
        c.font = copy(sc.font)
        c.border = copy(sc.border)
        c.fill = copy(sc.fill)
        c.number_format = copy(sc.number_format)
        c.protection = copy(sc.protection)
        c.alignment = copy(sc.alignment)

    def append_lines(self, sheetname, listing, headers=0):
        """
        Append multiple lines to the `sheetname` sheet

        `listing` must be a list of dict using column name (slugified) or column letter as position

        `headers` is forwarded to get_column_names method
        """
        for data in listing:
            self.append_line(sheetname, data, headers=headers)

    def append_line(self, sheetname, data, headers=0):
        """
        Append a new line to the `sheetname` sheet

        `data` must be a dict using column name (slugified) or column letter as position

        `headers` is forwarded to get_column_names method
        """

        assert isinstance(data, (dict, )), '{data} is not a dict'

        # Detect kind of mode to use
        count_letters = len(list(filter(lambda colname: re.search(r'^[A-Z]+$', colname), data.keys())))
        mode = 'letter' if count_letters == len(data.keys()) else 'name'

        # Check object keys
        headers_names = None
        if mode == 'name':
            headers_names = self.get_column_names(sheetname, headers=headers)
            invalid_keys = list(filter(lambda colname: colname not in headers_names, data.keys()))
            assert len(invalid_keys) == 0, f'Some keys are not supported in templated sheet: {", ".join(invalid_keys)}'

        # Copy line
        ws = self.wb.get_sheet_by_name(sheetname)
        tpl = self.get_template_line(sheetname, headers=headers)
        if not headers_names:
            headers_names = [get_column_letter(i) for i in range(len(tpl))]
        row = ws.max_row + 1
        for cell, cl in enumerate(headers_names):
            self.cell_copy(sheetname, row, (cell + 1), headers, data.get(cl, None))
